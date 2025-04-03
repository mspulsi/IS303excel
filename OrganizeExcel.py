# Markus Walker, Saxon Cullimore, Logan Hodges, Matt Pulsipher, Scott Peterson
# IS 303 Section 004
# This code organizes data from an original excel sheet and create a new one named "formatted_grades.xlsx".
# In the new workbook, the data from the original is divided into different worksheets by class.
# In each worksheet, the original data is broken up by the underscore and then the Last Name, First Name, and Student ID are stored in columns A, B, and C respectively.
# The grade of each student is stored in the column D.
# In column F are the titles of each summary statistic, which are calculated and displayed in column G.
# Each column with data is given a bolded title and an auto_filter is placed over the first four columns.

# Import necessary classes from openpyxl
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Load unorganized workbook, create new workbook, and remove existing sheet in new workbook
oldWB = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
newWB = Workbook()
newWB.remove(newWB["Sheet"])

# Declare initial variables
currSheet = oldWB.active
currClassName = ""
prevClassName = ""
firstRow = True

# Iterate through the cells in the original sheet starting from row 2 until last row and until column 3
max_row = currSheet.max_row
for row in currSheet.iter_rows(min_row=2, max_row=max_row, max_col=3):
    for cell in row:
        if cell.col_idx == 1:
            currClassName = cell.value
            # If class name does not have a corresponding worksheet, create one
            if currClassName != prevClassName:
                newWB.create_sheet(currClassName)
            prevClassName = currClassName
        # Split string from original worksheet by underscore
        elif cell.col_idx == 2:
            studentInfo = cell.value.split('_')
        else:
            studentInfo.append(cell.value)
    # Append divided data into new workbook
    newWB[currClassName].append(studentInfo)

# Instantiate font class to make column headers bold
f1 = Font(bold=True)

# Iterate through sheets in new workbook
for sheetName in newWB.sheetnames:
    # Insert row and assign column headers
    newWB[sheetName].insert_rows(1)
    newWB[sheetName]['A1'] = "Last Name"
    newWB[sheetName]['B1'] = "First Name"
    newWB[sheetName]['C1'] = "Student ID"
    newWB[sheetName]['D1'] = "Grade"
    newWB[sheetName]['F1'] = "Summary Statistics"
    newWB[sheetName]['G1'] = "Value"

    # Format column headers so that they are bold
    for cell in newWB[sheetName][1]:
        cell.font = f1

    # Iterate through columns and adjust width according to length of string in column header
    cols = ["A", "B", "C", "D", "F", "G"]
    for iCol in range(6):
        newWB[sheetName].column_dimensions[cols[iCol]].width = (len(newWB[sheetName][cols[iCol] + "1"].value) + 5)

    # Create an auto_filter for the data
    newWB[sheetName].auto_filter.ref = "A1:D1"

    # Assign titles for summary statistics
    newWB[sheetName]['F2'] = "Highest Grade"
    newWB[sheetName]['F3'] = "Lowest Grade"
    newWB[sheetName]['F4'] = "Mean Grade"
    newWB[sheetName]['F5'] = "Median Grade"
    newWB[sheetName]['F6'] = "Number of Students"

    # Calculate summary statistics from the second row until the last used row
    iLastRow = newWB[sheetName].max_row
    newWB[sheetName]["G2"] = f"=MAX(D2:D{iLastRow})"
    newWB[sheetName]["G3"] = f"=MIN(D2:D{iLastRow})"
    newWB[sheetName]["G4"] = f"=AVERAGE(D2:D{iLastRow})"
    newWB[sheetName]["G5"] = f"=MEDIAN(D2:D{iLastRow})"
    newWB[sheetName]["G6"] = f"=COUNT(D2:D{iLastRow})"

# Save new workbook and close both workbooks
newWB.save('formatted_grades.xlsx')
newWB.close()
oldWB.close()